# Authors: Joseph Rather, Chad Laursen, Elias Fobert, Isaac Johnson, Alayna Smith

# Discription: The organizes a messy dataset, splits it into clean and structured sheets by subject, 
# adds summary statistics, and applies visual formatting for clarity.

#Import Libraries
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook

#Import the poorly organized data
dataWorkbook = openpyxl.load_workbook('Poorly_Organized_Data_1.xlsx')
sheet = dataWorkbook.active

#Create a new workbook
myWorkbook = Workbook()
myWorkbook.remove(myWorkbook.active)

# create bold style
bold = Font(bold=True)

# make sheets with correct names
for row in sheet.iter_rows(min_row=2):
    if row[0].value not in myWorkbook.sheetnames:
        myWorkbook.create_sheet(row[0].value)

# Label column titles
for sheet_name in myWorkbook.sheetnames: 
    currSheet = myWorkbook[sheet_name]  
    currSheet["A1"] = "Last Name"
    currSheet["B1"] = "First Name"
    currSheet["C1"] = "Student ID"
    currSheet["D1"] = "Grade"

# go through unformatted data and put in corresponding sheet
for row in sheet.iter_rows(min_row=2, values_only= True):
    subject = row[0]
    myWorkbook.active = myWorkbook[subject]
    currSheet = myWorkbook.active
    values = row[1].split('_')
    grade = row[2]
    values.append(grade)
    currSheet.append(values)

# loop through sheets and change formatting and add necessary stats
for sheet_name in myWorkbook.sheetnames: 
    currSheet = myWorkbook[sheet_name]
    last_row = myWorkbook.active.max_row
    currSheet.auto_filter.ref = f"A1:D{last_row}"
    
    currSheet["F1"] = "Summary Statistics"
    currSheet["G1"] = "Value"
    currSheet["F2"] = "Highest Grade"
    currSheet["F3"] = "Lowest Grade"
    currSheet["F4"] = "Mean Grade"
    currSheet["F5"] = "Median Grade"
    currSheet["F6"] = "Number of Students"
    currSheet["G2"] = f"=MAX(D2:D{last_row})"
    currSheet["G3"] = f"=MIN(D2:D{last_row})"
    currSheet["G4"] = f"=AVERAGE(D2:D{last_row})"
    currSheet["G5"] = f"=MEDIAN(D2:D{last_row})"
    currSheet["G6"] = f"=COUNT(D2:D{last_row})"
    
    for cell in sum(currSheet["A1:G1"],()):
        cell.font = bold
        num_characters = len(str(cell.value))
        adj_width = num_characters + 5
        currSheet.column_dimensions[cell.column_letter].width = adj_width


# save and close 
myWorkbook.save(filename="formatted_grades.xlsx")
myWorkbook.close()